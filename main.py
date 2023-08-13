import os
import xmltodict
from datetime import datetime
from openpyxl import Workbook, load_workbook

# Diretório onde os arquivos XML estão localizados
diretorio_xml = "C:\\caminho\\para\\seus\\arquivos\\XML"

# Diretório onde serão salvos os arquivos Excel
diretorio_excel = "C:\\caminho\\para\\diretorio\\Excel"

# Obter a data atual
data_atual = datetime.now().strftime("%Y-%m-%d")

# Verificar se já existe um arquivo Excel para a data atual
nome_arquivo_excel = os.path.join(diretorio_excel, f"notas_fiscais_{data_atual}.xlsx")
if not os.path.exists(nome_arquivo_excel):
    # Criar um novo arquivo Excel se ele ainda não existir
    wb = Workbook()
    ws = wb.active
    headers = ["Número da Nota Fiscal", "Data de Emissão", "Descrição", "Valor da Nota Fiscal", "Forma de Pagamento"]
    ws.append(headers)
else:
    # Carregar o arquivo Excel existente
    wb = load_workbook(nome_arquivo_excel)
    ws = wb.active

# Percorrer todos os arquivos XML no diretório
for arquivo in os.listdir(diretorio_xml):
    if arquivo.endswith(".xml"):
        caminho_arquivo = os.path.join(diretorio_xml, arquivo)
        
        # Ler o conteúdo do arquivo XML com a codificação utf-8
        with open(caminho_arquivo, "r", encoding="utf-8") as xml_file:
            xml_data = xml_file.read()
        
        # Converter o XML em um dicionário Python
        data_dict = xmltodict.parse(xml_data)
        
        # Extrair as informações desejadas
        nota_fiscal = data_dict['NFES']['NOTA_FISCAL']
        numero_nf = nota_fiscal['numero_nf']
        data_emissao = nota_fiscal['data_emissao']
        descricao = nota_fiscal['descricao']
        valor_nf = nota_fiscal['valor_nf']
        forma_de_pagamento = nota_fiscal['forma_de_pagamento']
        
        # Formatar a data para o formato "DD-MM-AAAA"
        data_emissao_formatada = datetime.strptime(data_emissao, "%d/%m/%Y").strftime("%d-%m-%Y")
        
        # Adicionar informações à planilha
        ws.append([numero_nf, data_emissao_formatada, descricao, valor_nf, forma_de_pagamento])

# Salvar o arquivo Excel
wb.save(nome_arquivo_excel)

print("Arquivo Excel atualizado ou gerado:", nome_arquivo_excel)
